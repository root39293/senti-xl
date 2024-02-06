import sys
import pandas as pd
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QFileDialog, QLabel, QProgressBar, QTextEdit
from PyQt5.QtCore import QThread, pyqtSignal, Qt
import qdarkstyle
from openai import OpenAI
import openpyxl

OPENAI_API_KEY = "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"
client = OpenAI(api_key=OPENAI_API_KEY)

def analyze_sentiment(text):
    try:
        completion = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": """
                #지시문
                너는 지금부터 안전점검 서비스의 주관식 응답내용 해석 기계이다. 안전점검 서비스의 주관식 응답내용 해석 기계의 역할로서 제약조건, 출력형태를 준수하여 사용자가 입력하는 내용에 대해 감정분석을 하는 것이 너의 역할이다. 예시를 참고하여 답하여라.
                
                #제약조건
                - 가급적 짧게 답한다.
                - 부연설명은 하지 않는다.
                - 점검원 칭찬, 방문지연, 불친절, 안내미흡, 가스누설, 점검만족 등의 영역에서 감정을 분석한다.
                
                #출력형태
                [ 사용자의 입력내용에 따른 감정 긍정, 부정 출력 ]
                
                #예시
                점검원이 매우 친절하고 세심하게 설명해주었습니다.
                긍정
                 
                가스 누설 위험을 빠르게 감지하고 적절한 조치를 취해줘서 감사합니다.
                긍정
                
                불친절한 태도와 부족한 설명으로 서비스에 대해 다소 불만족합니다.
                부정
                 
                점검원의 방문지연과 안내미흡에 대해 불만족스러웠습니다.
                부정
                 
                """},
                {"role": "user", "content": f'#사용자의 입력 {text}'}
            ]
        )
        sentiment = completion.choices[0].message.content.strip()
        return sentiment
    except Exception as e:
        print(f"Error: {e}")
        return "분석 오류"

class AnalysisThread(QThread):
    finished = pyqtSignal(str)
    progress = pyqtSignal(int)
    log = pyqtSignal(str)

    def __init__(self, fileName):
        super().__init__()
        self.fileName = fileName

    def run(self):
        df = pd.read_excel(self.fileName)
        df['번호'] = df['번호'].astype(str)
        df.dropna(subset=['주관식응답내용'], inplace=True)
        
        total = len(df['주관식응답내용'])
        sentiment_results = []
        for i, text in enumerate(df['주관식응답내용'], 1):
            sentiment = analyze_sentiment(text)
            sentiment_results.append(sentiment)
            progress = int((i / total) * 100)
            self.progress.emit(progress)
            self.log.emit(f"분석 중... {i}/{total}")
        
        df['응답내용분석결과'] = sentiment_results
        
        outputFileName = self.fileName.replace('.xlsx', '_감정분석결과.xlsx')
        df.to_excel(outputFileName, index=False, engine='openpyxl')
        workbook = openpyxl.load_workbook(outputFileName)
        worksheet = workbook.active

        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = 50

        for col in worksheet.iter_cols():
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            if col_letter == 'M':  
                worksheet.column_dimensions[col_letter].width = 40
            else: 
                worksheet.column_dimensions[col_letter].width = 10

        workbook.save(outputFileName)

        
        self.finished.emit(outputFileName)

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('감정 분석 프로그램')
        self.setGeometry(100, 100, 500, 400)
        layout = QVBoxLayout()

        self.label = QLabel("파일을 선택하여 감정 분석을 시작하세요.")
        self.label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.label)

        self.logEdit = QTextEdit()
        self.logEdit.setReadOnly(True)
        layout.addWidget(self.logEdit)

        self.progressBar = QProgressBar(self)
        self.progressBar.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.progressBar)

        self.buttonSelect = QPushButton('파일 선택')
        self.buttonSelect.clicked.connect(self.openFileDialog)
        layout.addWidget(self.buttonSelect)

        self.buttonRun = QPushButton('실행')
        self.buttonRun.clicked.connect(self.runAnalysis)
        self.buttonRun.setEnabled(False)
        layout.addWidget(self.buttonRun)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)
        
        self.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())

    def openFileDialog(self):
        options = QFileDialog.Options()
        fileName, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 선택", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if fileName:
            self.fileName = fileName
            self.label.setText(f"선택된 파일: {fileName}")
            self.logEdit.append("파일이 선택되었습니다.")
            self.buttonRun.setEnabled(True)  

    def runAnalysis(self):
        self.logEdit.append("분석을 시작합니다...")
        self.analyzeFile(self.fileName)

    def analyzeFile(self, fileName):
        self.thread = AnalysisThread(fileName)
        self.thread.progress.connect(self.progressBar.setValue)
        self.thread.log.connect(self.logEdit.append)
        self.thread.finished.connect(self.onFinished)
        self.thread.start()

    def onFinished(self, outputFileName):
        self.label.setText(f"분석 완료! 결과 파일: {outputFileName}")
        self.logEdit.append(f"분석 완료! 결과 파일: {outputFileName}")
        self.progressBar.setValue(100)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())  
    mainWindow = MainWindow()
    mainWindow.show()
    sys.exit(app.exec_())